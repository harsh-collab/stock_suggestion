import pandas as pd
from bs4 import BeautifulSoup
import requests
import xlsxwriter

url = "https://chartink.com/screener/process"
l1 = []
# rule = {"scan_clause" : "( {cash} ( [0] 5 minute rsi  ( 14 ) < 30 and [0] 5 minute ema ( [0] 5 minute close , 100 ) > [0] 5 minute close ) ) )" }

rule = {"scan_clause" : "( {cash} ( latest rsi( 14 ) > latest ema( latest rsi( 9 ) , 21 ) and 1 day ago  rsi( 14 ) <= 1 day ago  ema( latest rsi( 9 ) , 21 ) and latest rsi( 14 ) > 70 and market cap > 1500 ) ) "}


with requests.session() as s:
    my_data = s.get(url)
    soup =  BeautifulSoup(my_data.content)
    
    meta = soup.find("meta", {"name" :"csrf-token"})["content"]
    
    header = {"x-csrf-token" : meta}
    data = s.post(url , headers=header , data = rule).json()
    
    stock_list = pd.DataFrame(data["data"])
    
    l1.append(stock_list["nsecode"][0])
    
df = pd.DataFrame({"stocks": l1})
df.to_excel("C://Users//HARSH//OneDrive//Desktop//stock50.xlsx",index=False)
print(l1)





import openpyxl
import webbrowser
import time

webbrowser.register('chrome',None,webbrowser.BackgroundBrowser("C://Program Files//Google//Chrome//Application//chrome.exe"))

wb1 = openpyxl.load_workbook('C://Users//HARSH//OneDrive//Desktop//all_instruments.xlsx')
sheet1 = wb1.active

wb2 = openpyxl.load_workbook('C://Users//HARSH//OneDrive//Desktop//stock50.xlsx')
sheet2 = wb2.active

x = 0
for j in range(1,100):
    for i in range(1+x,10+x):
        stock = sheet2.cell(row=i, column=1)
        for row in sheet1.rows:
            if stock.value == row[1].value:
                api = row[0].value
#                 print(stock.value,api)
                webbrowser.get('chrome').open('https://kite.zerodha.com/chart/ext/ciq/NSE/'+stock.value+'/'+str(api))
    x+=10
    time.sleep(40)
    user = input('do you want to see next charts (y/n): ')
    if user == 'y':
        continue
    else:
        break
